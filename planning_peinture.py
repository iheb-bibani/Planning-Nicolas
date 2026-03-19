"""
planning_peinture.py
────────────────────
Extracteur automatique du Planning Peinture — Solution complète en un seul fichier.

Installation :
    pip install streamlit openpyxl pandas

Lancement :
    streamlit run planning_peinture.py
"""

import io
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ══════════════════════════════════════════════════════════════════════════════
# 1. CONFIGURATION PAR DÉFAUT
# ══════════════════════════════════════════════════════════════════════════════

DEFAULT_CONFIG = {
    "codes_debut":     ["WB AI", "WB ST", "LR ST", "SA ST", "SA AI", "WBST", "LRST", "SAST", "WBAI"],
    "codes_fin":       ["CLT", "AS"],   # CLT prioritaire sur AS
    "seuil_msn":       10000,           # MSI/MSE = MSN < ce seuil
    "marqueur_interne":  "SALLES INTERNES",
    "marqueur_externe":  "SALLES EXTERNES",
    "marqueur_fin_zone": "MANAGEMENT",
    "marqueur_expleo":   "EXPLEO",
}


# ══════════════════════════════════════════════════════════════════════════════
# 2. STRUCTURES DE DONNÉES
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class Cycle:
    msn:        str
    debut:      Optional[datetime]
    fin:        Optional[datetime]
    salle:      str
    zone:       str           # "interne" ou "externe"
    is_reprise: bool = False
    is_expleo:  bool = False

    @property
    def msn_int(self) -> int:
        try:    return int(self.msn)
        except: return 0

    @property
    def debut_str(self) -> str:
        return self.debut.strftime("%d/%m/%Y") if self.debut else "À vérifier"

    @property
    def fin_str(self) -> str:
        return self.fin.strftime("%d/%m/%Y") if self.fin else "À vérifier"


@dataclass
class Zones:
    interne_debut: int
    interne_fin:   int
    externe_debut: int
    externe_fin:   int
    expleo_debut:  int


@dataclass
class Resultat:
    SI:   list = field(default_factory=list)
    MSI:  list = field(default_factory=list)
    SE:   list = field(default_factory=list)
    MSE:  list = field(default_factory=list)
    logs: list = field(default_factory=list)  # [(niveau, message)]

    def log(self, niveau, msg): self.logs.append((niveau, msg))

    @property
    def total(self): return len(self.SI) + len(self.MSI) + len(self.SE) + len(self.MSE)


# ══════════════════════════════════════════════════════════════════════════════
# 3. UTILITAIRES EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def cell_val(ws, row, col) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""

def cell_bg(ws, row, col) -> Optional[str]:
    cell = ws.cell(row=row, column=col)
    if cell.fill and cell.fill.fill_type == "solid":
        c = cell.fill.fgColor
        if c and c.type == "rgb":
            return c.rgb.upper()
    return None

def is_reprise(ws, row, col) -> bool:
    return cell_bg(ws, row, col) == "FFFFFF00"

def build_date_map(ws) -> dict:
    """Ligne 5 = dates du planning."""
    return {
        col: ws.cell(row=5, column=col).value
        for col in range(1, ws.max_column + 1)
        if isinstance(ws.cell(row=5, column=col).value, datetime)
    }

def get_date(col, date_map) -> Optional[datetime]:
    if col in date_map: return date_map[col]
    if not date_map:    return None
    return date_map[min(date_map, key=lambda c: abs(c - col))]

def build_merge_map(ws) -> dict:
    m = {}
    for r in ws.merged_cells.ranges:
        for row in range(r.min_row, r.max_row + 1):
            for col in range(r.min_col, r.max_col + 1):
                m[(row, col)] = (r.min_col, r.max_col)
    return m

def extraire_msn(texte) -> Optional[str]:
    for p in texte.split():
        if p.isdigit() and 3 <= len(p) <= 6:
            return p
    return None

def extraire_salle_ext(texte) -> str:
    parts = texte.split()
    return parts[0][:3] if parts else "?"


# ══════════════════════════════════════════════════════════════════════════════
# 4. DÉTECTION DES ZONES
# ══════════════════════════════════════════════════════════════════════════════

def detect_zones(ws, cfg) -> Zones:
    interne_debut = externe_debut = expleo_debut = None
    fin_zone = ws.max_row

    for row in range(1, ws.max_row + 1):
        val = cell_val(ws, row, 1).upper()
        if not val: continue

        if interne_debut is None and cfg["marqueur_interne"] in val:
            interne_debut = row + 1
        if externe_debut is None and cfg["marqueur_externe"] in val:
            externe_debut = row + 1
        if expleo_debut is None and externe_debut and cfg["marqueur_expleo"] in val:
            expleo_debut = row
        if cfg["marqueur_fin_zone"] in val:
            fin_zone = row - 1
            break

    return Zones(
        interne_debut = interne_debut or 7,
        interne_fin   = (externe_debut - 2) if externe_debut else fin_zone,
        externe_debut = externe_debut or 78,
        externe_fin   = fin_zone,
        expleo_debut  = expleo_debut or externe_debut or 78,
    )


# ══════════════════════════════════════════════════════════════════════════════
# 5. EXTRACTION DES CYCLES
# ══════════════════════════════════════════════════════════════════════════════

def chercher_fin(ws, row, col_debut, cfg, date_map) -> Optional[datetime]:
    codes_fin   = [c.upper() for c in cfg["codes_fin"]]
    codes_debut = [c.upper() for c in cfg["codes_debut"]]
    max_col     = ws.max_column

    for col in range(col_debut + 1, min(col_debut + 100, max_col + 1)):
        val = cell_val(ws, row, col).upper()
        if not val: continue
        for code in codes_fin:
            if val.startswith(code):
                return get_date(col, date_map)
        if any(val.startswith(c) for c in codes_debut) and col > col_debut + 3:
            break
    return None


def extraire_internes(ws, zones, date_map, merge_map, cfg) -> list:
    cycles      = []
    codes_debut = [c.upper() for c in cfg["codes_debut"]]
    max_col     = ws.max_column

    for row in range(zones.interne_debut, zones.interne_fin + 1):
        salle = cell_val(ws, row, 1)
        # Lignes de salles : "C32", "C33", "C35", "C39"...
        if not salle or not (salle.startswith("C") and salle[1:].isdigit()):
            continue

        col = 5
        while col <= max_col:
            val = cell_val(ws, row, col)
            if not val: col += 1; continue

            # ── Reprise (fond jaune) ──────────────────────────────────────
            # Début = date de la 1ère colonne de la cellule fusionnée
            # Fin   = date de la dernière colonne de la cellule fusionnée
            if is_reprise(ws, row, col):
                mg        = merge_map.get((row, col))
                start_col = mg[0] if mg else col
                end_col   = mg[1] if mg else col
                msn = extraire_msn(val)
                if msn:
                    cycles.append(Cycle(
                        msn=msn, debut=get_date(start_col, date_map),
                        fin=get_date(end_col, date_map),
                        salle=salle, zone="interne", is_reprise=True
                    ))
                col = end_col + 1; continue

            # ── Début de cycle normal ─────────────────────────────────────
            if any(val.upper().startswith(c) for c in codes_debut):
                msn_cell = ""
                for nc in range(col + 1, min(col + 5, max_col + 1)):
                    nv = cell_val(ws, row, nc)
                    if nv and not any(nv.upper().startswith(c) for c in codes_debut):
                        msn_cell = nv; break
                msn = extraire_msn(msn_cell)
                if msn:
                    cycles.append(Cycle(
                        msn=msn, debut=get_date(col, date_map),
                        fin=chercher_fin(ws, row, col, cfg, date_map),
                        salle=salle, zone="interne"
                    ))
            col += 1

    return cycles


def extraire_externes(ws, zones, date_map, merge_map, cfg) -> list:
    cycles      = []
    codes_debut = [c.upper() for c in cfg["codes_debut"]]
    max_col     = ws.max_column

    for row in range(zones.externe_debut, zones.externe_fin + 1):
        col4 = cell_val(ws, row, 4).upper()
        if not (col4.startswith("ACTIVITE") or col4 == "POSITION" or col4 == ""):
            continue
        is_expleo = row >= zones.expleo_debut

        col = 5
        while col <= max_col:
            val = cell_val(ws, row, col)
            if not val: col += 1; continue

            # ── Reprise (fond jaune) ──────────────────────────────────────
            # Début = date de la 1ère colonne de la cellule fusionnée
            # Fin   = date de la dernière colonne de la cellule fusionnée
            if is_reprise(ws, row, col):
                mg        = merge_map.get((row, col))
                start_col = mg[0] if mg else col
                end_col   = mg[1] if mg else col
                msn   = extraire_msn(val)
                salle = extraire_salle_ext(val)
                if msn:
                    cycles.append(Cycle(
                        msn=msn, debut=get_date(start_col, date_map),
                        fin=get_date(end_col, date_map),
                        salle=salle, zone="externe",
                        is_reprise=True, is_expleo=is_expleo
                    ))
                col = end_col + 1; continue

            # ── Début de cycle normal ─────────────────────────────────────
            if any(val.upper().startswith(c) for c in codes_debut):
                msn_cell = ""
                for nc in range(col + 1, min(col + 5, max_col + 1)):
                    nv = cell_val(ws, row, nc)
                    if nv and not any(nv.upper().startswith(c) for c in codes_debut):
                        msn_cell = nv; break
                # Ignorer les doublons BINOME
                if "BINOME" in msn_cell.upper():
                    col += 1; continue
                msn   = extraire_msn(msn_cell)
                salle = extraire_salle_ext(msn_cell) if msn_cell else "?"
                if msn or msn_cell:
                    cycles.append(Cycle(
                        msn=msn or msn_cell, debut=get_date(col, date_map),
                        fin=chercher_fin(ws, row, col, cfg, date_map),
                        salle=salle, zone="externe", is_expleo=is_expleo
                    ))
            col += 1

    return cycles


# ══════════════════════════════════════════════════════════════════════════════
# 6. CLASSIFICATION SI / MSI / SE / MSE
# ══════════════════════════════════════════════════════════════════════════════

def classifier(cycles, cfg, resultat):
    """
    SI  = tous les internes
    MSI = internes avec MSN < seuil
    SE  = tous les EXPLEO
    MSE = externes avec MSN < seuil
    """
    seuil = cfg["seuil_msn"]
    for c in cycles:
        petit = 0 < c.msn_int < seuil
        if c.zone == "interne":
            resultat.SI.append(c)
            if petit: resultat.MSI.append(c)
        else:
            if c.is_expleo: resultat.SE.append(c)
            if petit:       resultat.MSE.append(c)


# ══════════════════════════════════════════════════════════════════════════════
# 7. FONCTION PRINCIPALE D'EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def extraire(fichier, cfg) -> Resultat:
    res = Resultat()
    try:
        wb = openpyxl.load_workbook(fichier, data_only=False)
    except Exception as e:
        res.log("erreur", f"Impossible d'ouvrir le fichier : {e}"); return res

    if "Planning" not in wb.sheetnames:
        res.log("erreur", 'Feuille "Planning" introuvable.'); return res

    ws = wb["Planning"]
    res.log("ok",   f"Feuille 'Planning' chargée — {ws.max_row} lignes × {ws.max_column} colonnes")

    date_map  = build_date_map(ws)
    merge_map = build_merge_map(ws)
    res.log("info", f"{len(date_map)} colonnes de dates · {len(merge_map)} cellules fusionnées")

    zones = detect_zones(ws, cfg)
    res.log("ok",   f"Zone INTERNE  : lignes {zones.interne_debut} → {zones.interne_fin}")
    res.log("ok",   f"Zone EXTERNE  : lignes {zones.externe_debut} → {zones.externe_fin}")
    res.log("ok",   f"Sous-zone SE (EXPLEO) : à partir de la ligne {zones.expleo_debut}")

    cycles_int = extraire_internes(ws, zones, date_map, merge_map, cfg)
    cycles_ext = extraire_externes(ws, zones, date_map, merge_map, cfg)
    tous       = cycles_int + cycles_ext

    res.log("info", f"{len(cycles_int)} cycles internes · {len(cycles_ext)} cycles externes")

    classifier(tous, cfg, res)

    res.log("ok",   f"SI={len(res.SI)}  MSI={len(res.MSI)}  SE={len(res.SE)}  MSE={len(res.MSE)}")

    sans_fin  = [c for c in tous if c.fin is None]
    reprises  = [c for c in tous if c.is_reprise]
    if sans_fin: res.log("attention", f"{len(sans_fin)} cycle(s) sans date de fin — à vérifier manuellement")
    if reprises: res.log("info",      f"{len(reprises)} reprise(s) détectée(s)")

    res.log("ok", "Extraction terminée ✓")
    return res


# ══════════════════════════════════════════════════════════════════════════════
# 8. EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════════════

COULEURS_ZONE = {
    "SI":  {"bg": "DBEAFE", "txt": "1D4ED8"},
    "MSI": {"bg": "EDE9FE", "txt": "6D28D9"},
    "SE":  {"bg": "D1FAE5", "txt": "065F46"},
    "MSE": {"bg": "FEF3C7", "txt": "92400E"},
}

def exporter_excel(fichier_bytes, res: Resultat) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin   = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for zone in ["SI", "MSI", "SE", "MSE"]:
        cycles = getattr(res, zone)
        ws     = wb.create_sheet(title=zone)
        col    = COULEURS_ZONE[zone]

        # Titre zone
        ws.merge_cells("A1:E1")
        ws["A1"] = f"Extraction automatique — {zone}"
        ws["A1"].font      = Font(name="Calibri", bold=True, size=13, color=col["txt"])
        ws["A1"].fill      = PatternFill("solid", fgColor=col["bg"])
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 28

        # En-têtes
        for ci, h in enumerate(["MSN", "Début", "Fin", "Salle", "Type"], 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="1E293B")
            c.alignment = center
            c.border    = border
        ws.row_dimensions[3].height = 22

        # Données
        for ri, cycle in enumerate(cycles, 4):
            alt  = PatternFill("solid", fgColor="F8FAFC" if ri % 2 == 0 else "FFFFFF")
            vals = [cycle.msn, cycle.debut, cycle.fin, cycle.salle,
                    "Reprise" if cycle.is_reprise else "Cycle"]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font      = Font(name="Calibri", size=10)
                c.fill      = alt
                c.border    = border
                c.alignment = Alignment(horizontal="left" if ci == 1 else "center",
                                        vertical="center")
                if isinstance(val, datetime):
                    c.number_format = "DD/MM/YYYY"
                if val is None:
                    c.value = "À vérifier"
                    c.font  = Font(name="Calibri", size=10, italic=True, color="EF4444")
                if ci == 5 and cycle.is_reprise:
                    c.font = Font(name="Calibri", size=10, bold=True, color="92400E")
                    c.fill = PatternFill("solid", fgColor="FEF3C7")

        # Largeurs
        for letter, w in zip("ABCDE", [14, 16, 16, 10, 12]):
            ws.column_dimensions[letter].width = w
        ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# 9. INTERFACE STREAMLIT
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Planning Peinture", page_icon="✈️", layout="wide")

st.markdown("""
<style>
    #MainMenu, footer { visibility: hidden; }
    .block-container { padding-top: 2rem; }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("# ✈️ Planning Peinture — Extracteur automatique")
st.markdown("Déposez votre fichier Planning pour remplir automatiquement les onglets **SI · MSI · SE · MSE**")
st.divider()

# ── Sidebar : configuration ───────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Configuration")
    st.divider()

    codes_debut = st.text_area(
        "Codes début de cycle (un par ligne)",
        value="\n".join(DEFAULT_CONFIG["codes_debut"]), height=175
    )
    codes_fin = st.text_area(
        "Codes fin de cycle (CLT en premier = prioritaire)",
        value="\n".join(DEFAULT_CONFIG["codes_fin"]), height=75
    )
    st.divider()

    seuil = st.number_input(
        "Seuil MSN pour MSI / MSE (MSN strictement inférieur à)",
        min_value=100, max_value=99999,
        value=DEFAULT_CONFIG["seuil_msn"], step=1000
    )
    st.divider()

    st.markdown("**Marqueurs de zones (colonne A)**")
    m_int = st.text_input("Début zone interne",  DEFAULT_CONFIG["marqueur_interne"])
    m_ext = st.text_input("Début zone externe",  DEFAULT_CONFIG["marqueur_externe"])
    m_fin = st.text_input("Fin de toutes zones", DEFAULT_CONFIG["marqueur_fin_zone"])

cfg = {
    "codes_debut":      [c.strip().upper() for c in codes_debut.splitlines() if c.strip()],
    "codes_fin":        [c.strip().upper() for c in codes_fin.splitlines()   if c.strip()],
    "seuil_msn":        seuil,
    "marqueur_interne":  m_int.strip().upper(),
    "marqueur_externe":  m_ext.strip().upper(),
    "marqueur_fin_zone": m_fin.strip().upper(),
    "marqueur_expleo":   DEFAULT_CONFIG["marqueur_expleo"],
}

# ── Upload ────────────────────────────────────────────────────────────────────
col_up, col_rules = st.columns([2, 1])

with col_up:
    fichier = st.file_uploader("📂 Fichier Planning (.xlsx)", type=["xlsx"])

with col_rules:
    st.info("""
    **Règles de classification**
    - 🔵 **SI** — Tous les cycles internes
    - 🟣 **MSI** — Internes, MSN < seuil
    - 🟢 **SE** — Tous les cycles EXPLEO
    - 🟡 **MSE** — Externes, MSN < seuil
    """)

if not fichier:
    st.stop()

# ── Extraction ────────────────────────────────────────────────────────────────
fichier_bytes = fichier.read()

with st.spinner("🔍 Analyse en cours..."):
    res = extraire(io.BytesIO(fichier_bytes), cfg)

# ── Logs ──────────────────────────────────────────────────────────────────────
ICONES = {"ok": "✅", "info": "ℹ️", "attention": "⚠️", "erreur": "❌"}
with st.expander("📋 Journal d'extraction"):
    for niveau, msg in res.logs:
        st.markdown(f"{ICONES.get(niveau, '•')} `{msg}`")

if any(n == "erreur" for n, _ in res.logs):
    st.stop()

# ── Métriques ─────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("### 📊 Résumé")
c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("🔵 SI",  len(res.SI))
c2.metric("🟣 MSI", len(res.MSI))
c3.metric("🟢 SE",  len(res.SE))
c4.metric("🟡 MSE", len(res.MSE))
c5.metric("Total",  res.total)

# ── Résultats ─────────────────────────────────────────────────────────────────
st.markdown("### 📋 Détail des cycles")
tabs = st.tabs(["🔵 SI", "🟣 MSI", "🟢 SE", "🟡 MSE"])

for tab, zone in zip(tabs, ["SI", "MSI", "SE", "MSE"]):
    with tab:
        cycles = getattr(res, zone)
        nb_rep = sum(1 for c in cycles if c.is_reprise)
        nb_ko  = sum(1 for c in cycles if c.fin is None)

        m1, m2, m3 = st.columns(3)
        m1.metric("Total", len(cycles))
        m2.metric("Reprises", nb_rep)
        m3.metric("⚠️ Sans date de fin", nb_ko)

        if cycles:
            df = pd.DataFrame([{
                "MSN":   str(c.msn),
                "Début": c.debut_str,
                "Fin":   c.fin_str,
                "Salle": str(c.salle),
                "Type":  "Reprise" if c.is_reprise else "Cycle",
            } for c in cycles])
            # Forcer tous les types en string — évite l'erreur LargeUtf8
            df = df.astype(str)
            st.dataframe(df, use_container_width=True)
        else:
            st.info("Aucun cycle détecté pour cette zone.")

# ── Téléchargement ────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("### ⬇️ Télécharger le résultat")

with st.spinner("Génération du fichier Excel..."):
    excel_bytes = exporter_excel(fichier_bytes, res)

st.download_button(
    label="⬇️ Télécharger le fichier Excel complété",
    data=excel_bytes,
    file_name=fichier.name.replace(".xlsx", "_extrait.xlsx"),
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
    type="primary",
)
